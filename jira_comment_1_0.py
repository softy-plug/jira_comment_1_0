import os

os.system("pip install jira")
os.system("pip install openpyxl")

# import jira
import openpyxl
from openpyxl import load_workbook
from jira import JIRA

input("Для запуска программы нажмите Enter")

exl_path = r"jira.xlsx"
exl = load_workbook(exl_path)  # Запись пути к таблице

# работа с листом data

sheet_data = exl["data"]  # назначить лист data в переменную

login_exl = sheet_data.cell(row=2, column=1).value  # добавить в переменную логин из таблицы
password_exl = sheet_data.cell(row=2, column=2).value  # добавить в переменную пароль из таблицы
# server_exl = sheet_data.cell(row=2, column=4).value  # добавить в переменную сервер из таблицы

# работа с листом jira

sheet_jr = exl["jr"] # назначить лист main в переменную

# Iterate over all rows in the specified column and add comment to each issue
for row in range(2, sheet_jr.max_row + 1):  # Start from row 2
    # Read issue key and comment from the current row
    jr_exl = sheet_jr.cell(row=row, column=1).value
    jr_comment = sheet_jr.cell(row=row, column=2).value

    # Connect to JIRA server
    jira = JIRA(server='https://jr.synergy.ru', basic_auth=(login_exl, password_exl))

    # JIRA issue key
    issue_key = jr_exl

    # Comment to add
    comment = jr_comment

    # Add comment to the issue
    jira.add_comment(issue_key, comment)

input("Отправка сообщений в заявки завершена. Нажмите Enter для закрытия окна")

# softy_plug